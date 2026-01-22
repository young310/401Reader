# excel_export_service.py
# Excel 匯出服務

from io import BytesIO
from typing import Generator, List, Dict, Any
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
import os

from app.models import TaxOcrJob as Job, TaxOcrVersion


# ==================== 輔助函數 ====================

def add_title_rows(ws, company_name: str, doc_type: str, fiscal_year: int, total_cols: int) -> int:
    """
    新增標題行（前3行 + 空行）
    
    Returns:
        下一行的行號
    """
    # 第1行：公司名稱
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    ws['A1'] = company_name or "○○○股份有限公司"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 第2行：申報書類型
    ws.merge_cells(f'A2:{get_column_letter(total_cols)}2')
    ws['A2'] = doc_type
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 第3行：年度
    ws.merge_cells(f'A3:{get_column_letter(total_cols)}3')
    ws['A3'] = f"{fiscal_year}年度" if fiscal_year else "年度"
    ws['A3'].font = Font(size=12, bold=True)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 第4行：空行
    ws.row_dimensions[4].height = 15
    
    return 5  # 返回下一行的行號


def add_section_header(ws, row: int, title: str, total_cols: int, bg_color: str) -> int:
    """
    新增區段標題（銷項/進項）
    
    Returns:
        下一行的行號
    """
    ws.merge_cells(f'A{row}:{get_column_letter(total_cols)}{row}')
    cell = ws[f'A{row}']
    cell.value = title
    cell.font = Font(size=12, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    # 加上黑框
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    cell.border = thin_border
    return row + 1


def add_subtotal_border(ws, row: int, start_col: int, end_col: int):
    """為項目總額列新增邊框（上單線，下雙線）- 只在指定欄位"""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(
            top=Side(style='thin'),
            bottom=Side(style='double')
        )


def add_total_border(ws, row: int, start_col: int, end_col: int):
    """為合計列新增邊框（上單線，下雙線）"""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(
            top=Side(style='thin'),
            bottom=Side(style='double')
        )


def set_column_widths(ws, widths: List[int]):
    """設定欄寬"""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def format_number_cells(ws, row: int, start_col: int, end_col: int):
    """設定數字格式（會計格式，有千分位）"""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1


def add_black_borders(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    """為指定範圍的儲存格加上黑框"""
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border


def format_page_number(record: dict) -> str:
    """
    格式化頁碼顯示

    若有憑單序號（股利憑單），顯示為 "頁碼-序號" 格式（如 1-1, 1-2）
    否則只顯示頁碼

    Args:
        record: 包含頁碼和憑單序號的記錄

    Returns:
        格式化後的頁碼字串
    """
    page_num = record.get("頁碼", "")
    voucher_seq = record.get("憑單序號")

    # 如果頁碼為空或 None，返回空字串
    if page_num is None or page_num == "":
        return ""

    # 如果有憑單序號，顯示為 "頁碼-序號" 格式
    if voucher_seq is not None and voucher_seq != "":
        return f"{page_num}-{voucher_seq}"

    # 否則只顯示頁碼
    return str(page_num)


def create_excel_export_stream(
    jobs: Generator[Job, None, None],
    stream: str
) -> BytesIO:
    """
    根據 Jobs 生成 Excel 檔案

    Args:
        jobs: Job 物件的生成器（用於串流處理，避免記憶體溢出）
        stream: "支出" 或 "收入"

    Returns:
        Excel 檔案的 BytesIO 物件

    Note:
        目前為佔位實作，等待使用者提供明確的 Excel 模板後填充邏輯
    """
    # 建立空的 Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{stream}彙總表"

    # TODO: 根據使用者提供的 Excel 模板填充資料
    # 目前僅建立一個簡單的示範表格

    # 表頭
    ws.append(["檔案名稱", "文件類型", "處理時間", "資料摘要"])

    # 逐筆寫入資料（串流處理）
    for job in jobs:
        ws.append([
            job.original_filename,
            job.get_display_document_type(),  # 使用新的顯示方法
            job.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
            f"JSON 資料長度: {len(str(job.result_json))}"
        ])

    # 儲存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def aggregate_type1_results(jobs: list[Job]) -> dict:
    """
    聚合 TYPE1 (401/403) 的結果

    Args:
        jobs: Job 物件列表

    Returns:
        聚合後的資料

    Note:
        TYPE1 每頁獨立，不需要聚合，直接轉換為陣列
    """
    results = []
    for job in jobs:
        if job.result_json:
            results.append(job.result_json)
    return {"資料": results}


def aggregate_type2_results(jobs: list[Job]) -> dict:
    """
    聚合 TYPE2 (彙總表) 的結果

    Args:
        jobs: Job 物件列表

    Returns:
        聚合後的資料

    Note:
        TYPE2 每頁獨立，不需要聚合，直接轉換為陣列
    """
    results = []
    for job in jobs:
        if job.result_json:
            results.append(job.result_json)
    return {"資料": results}


def create_excel_from_version(version: TaxOcrVersion) -> BytesIO:
    """
    根據版本資料生成 Excel 檔案

    Args:
        version: TaxOcrVersion 物件

    Returns:
        Excel 檔案的 BytesIO 物件
    """
    wb = Workbook()
    wb.remove(wb.active)  # 移除預設工作表

    # 根據 table_type 決定生成哪種 Excel
    if version.table_type in ['401', '403']:
        # 營業稅申報書
        create_401_excel(wb, version)
    elif version.table_type in ['withholding_income', 'withholding_expense', 'dividend_income', 'dividend_expense']:
        # 扣繳憑單（包括股利憑單）
        create_withholding_excel(wb, version)
    else:
        # 未知類型，建立簡單表格
        ws = wb.create_sheet("資料")
        ws.append(["未知表格類型", version.table_type])

    # 儲存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def create_401_excel(wb: Workbook, version: TaxOcrVersion):
    """
    建立 401/403 營業稅申報書 Excel

    格式：
    - 前3行：公司名稱、申報書類型、年度
    - 第4行：空行
    - 銷項區段（淺綠色標題）+ 3層表頭 + 資料 + 合計列
    - 空2行
    - 進項區段（淺藍色標題）+ 2層表頭 + 資料 + 合計列

    銷項表格 14 欄結構：
    月份 | 一般稅額銷售額(應稅:三聯式/二聯式, 零稅率:經海關/非經海關, 免稅) |
    特種稅額銷售額 | 其他 | 銷售額合計 | 銷貨退回及折讓(一般/零稅率/特種) |
    銷貨退回及折讓合計 | 淨額
    """
    ws = wb.create_sheet("營業稅申報書")

    # 設定欄寬（14欄）- 統一加寬
    # 月份 | 三聯式 | 二聯式 | 非經海關 | 經海關 | 免稅 | 特種稅額 | 其他 | 銷售額合計 | 退回-一般 | 退回-零稅率 | 退回-特種 | 退回合計 | 淨額
    column_widths = [12, 20, 20, 20, 20, 20, 20, 20, 18, 18, 18, 18, 22, 20]
    set_column_widths(ws, column_widths)

    # 新增標題行（前3行 + 空行）
    doc_type = "營業人銷售額與稅額申報書彙總表"
    current_row = add_title_rows(ws, version.company_name, doc_type, version.fiscal_year, len(column_widths))

    # === 銷項表格 ===

    # 銷項區段標題（更綠更亮的背景）
    current_row = add_section_header(ws, current_row, "銷項", len(column_widths), "CCFFCC")

    # 銷項表頭（3層）
    header_start_row = current_row

    # 第1層表頭
    ws.merge_cells(f'A{current_row}:A{current_row+2}')  # 月份 (rowspan=3)
    ws[f'A{current_row}'] = "月份"

    ws.merge_cells(f'B{current_row}:F{current_row}')    # 一般稅額銷售額 (colspan=5)
    ws[f'B{current_row}'] = "一般稅額銷售額"

    ws.merge_cells(f'G{current_row}:G{current_row+2}')  # 特種稅額銷售額 (rowspan=3)
    ws[f'G{current_row}'] = "特種稅額銷售額"

    ws.merge_cells(f'H{current_row}:H{current_row+2}')  # 其他 (rowspan=3)
    ws[f'H{current_row}'] = "其他"

    ws.merge_cells(f'I{current_row}:I{current_row+2}')  # 銷售額合計 (rowspan=3)
    ws[f'I{current_row}'] = "銷售額合計"

    ws.merge_cells(f'J{current_row}:L{current_row}')    # 銷貨退回及折讓 (colspan=3)
    ws[f'J{current_row}'] = "銷貨退回及折讓"

    ws.merge_cells(f'M{current_row}:M{current_row+2}')  # 銷貨退回及折讓合計 (rowspan=3)
    ws[f'M{current_row}'] = "銷貨退回及折讓合計"

    ws.merge_cells(f'N{current_row}:N{current_row+2}')  # 淨額 (rowspan=3)
    ws[f'N{current_row}'] = "淨額"

    current_row += 1

    # 第2層表頭
    ws.merge_cells(f'B{current_row}:C{current_row}')    # 應稅 (colspan=2)
    ws[f'B{current_row}'] = "應稅"

    ws.merge_cells(f'D{current_row}:E{current_row}')    # 零稅率 (colspan=2)
    ws[f'D{current_row}'] = "零稅率"

    ws.merge_cells(f'F{current_row}:F{current_row+1}')  # 免稅 (rowspan=2)
    ws[f'F{current_row}'] = "免稅"

    # 銷貨退回及折讓子欄（第2層延續到第3層）
    ws.merge_cells(f'J{current_row}:J{current_row+1}')  # 一般 (rowspan=2)
    ws[f'J{current_row}'] = "一般"

    ws.merge_cells(f'K{current_row}:K{current_row+1}')  # 零稅率 (rowspan=2)
    ws[f'K{current_row}'] = "零稅率"

    ws.merge_cells(f'L{current_row}:L{current_row+1}')  # 特種 (rowspan=2)
    ws[f'L{current_row}'] = "特種"

    current_row += 1

    # 第3層表頭
    ws[f'B{current_row}'] = "三聯式"
    ws[f'C{current_row}'] = "二聯式"
    ws[f'D{current_row}'] = "非經海關"
    ws[f'E{current_row}'] = "經海關"
    # F欄已被免稅佔用（rowspan=2）
    # J, K, L 欄已被銷貨退回子欄佔用（rowspan=2）

    # 設定表頭樣式
    for row in range(header_start_row, current_row + 1):
        for col in range(1, len(column_widths) + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:  # 只設定有值的儲存格
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row += 1

    # 填入銷項資料
    table_data = version.table_data or []
    sales_start_row = current_row
    sales_section_start = header_start_row - 1  # 包含「銷項」標題
    for record in table_data:
        ws.cell(row=current_row, column=1, value=record.get("month", ""))

        # 數字欄位 - 14 欄結構（非經海關在前，經海關在後）
        cells_data = [
            (2, record.get("triplicateSales", 0)),      # 三聯式
            (3, record.get("duplicateSales", 0)),       # 二聯式
            (4, record.get("nonCustomsSales", 0)),      # 非經海關
            (5, record.get("customsSales", 0)),         # 經海關
            (6, record.get("taxFreeSales", 0)),         # 免稅
            (7, record.get("specialTaxSales", 0)),      # 特種稅額銷售額
            (8, record.get("otherSales", 0)),           # 其他
            (9, record.get("totalSales", 0)),           # 銷售額合計
            (10, record.get("returnGeneral", 0)),       # 銷貨退回-一般
            (11, record.get("returnZeroRate", 0)),      # 銷貨退回-零稅率
            (12, record.get("returnSpecial", 0)),       # 銷貨退回-特種
            (13, record.get("returnTotal", 0)),         # 銷貨退回及折讓合計
            (14, record.get("netAmount", 0))            # 淨額
        ]

        for col, value in cells_data:
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1

        current_row += 1

    # 銷項合計列
    if table_data:
        ws.cell(row=current_row, column=1, value="合計")

        # 計算合計
        totals = {
            "triplicateSales": sum(r.get("triplicateSales", 0) for r in table_data),
            "duplicateSales": sum(r.get("duplicateSales", 0) for r in table_data),
            "taxFreeSales": sum(r.get("taxFreeSales", 0) for r in table_data),
            "customsSales": sum(r.get("customsSales", 0) for r in table_data),
            "nonCustomsSales": sum(r.get("nonCustomsSales", 0) for r in table_data),
            "specialTaxSales": sum(r.get("specialTaxSales", 0) for r in table_data),
            "otherSales": sum(r.get("otherSales", 0) for r in table_data),
            "totalSales": sum(r.get("totalSales", 0) for r in table_data),
            "returnGeneral": sum(r.get("returnGeneral", 0) for r in table_data),
            "returnZeroRate": sum(r.get("returnZeroRate", 0) for r in table_data),
            "returnSpecial": sum(r.get("returnSpecial", 0) for r in table_data),
            "returnTotal": sum(r.get("returnTotal", 0) for r in table_data),
            "netAmount": sum(r.get("netAmount", 0) for r in table_data)
        }

        total_cells_data = [
            (2, totals["triplicateSales"]),     # 三聯式
            (3, totals["duplicateSales"]),      # 二聯式
            (4, totals["nonCustomsSales"]),     # 非經海關
            (5, totals["customsSales"]),        # 經海關
            (6, totals["taxFreeSales"]),        # 免稅
            (7, totals["specialTaxSales"]),     # 特種稅額銷售額
            (8, totals["otherSales"]),          # 其他
            (9, totals["totalSales"]),          # 銷售額合計
            (10, totals["returnGeneral"]),      # 銷貨退回-一般
            (11, totals["returnZeroRate"]),     # 銷貨退回-零稅率
            (12, totals["returnSpecial"]),      # 銷貨退回-特種
            (13, totals["returnTotal"]),        # 銷貨退回及折讓合計
            (14, totals["netAmount"])           # 淨額
        ]

        for col, value in total_cells_data:
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
            cell.font = Font(bold=True)

        # 設定合計列邊框
        add_total_border(ws, current_row, 2, len(column_widths))

        sales_section_end = current_row
        current_row += 1
    else:
        sales_section_end = current_row - 1

    # 為銷項表格加上黑框（從「銷項」標題到合計列）
    add_black_borders(ws, sales_section_start, sales_section_end, 1, len(column_widths))
    
    # 空2行
    current_row += 2

    # === 進項表格 ===

    # 進項表格欄寬（8欄）- 特別加寬「進口免稅貨物金額」
    # 月份 | 得扣抵-進貨及費用 | 得扣抵-固定資產 | 退回-進貨及費用 | 退回-固定資產 | 合計-進貨及費用 | 合計-固定資產 | 進口免稅貨物金額
    purchase_column_widths = [12, 18, 18, 18, 18, 18, 18, 22]
    for i, width in enumerate(purchase_column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 進項區段標題（更亮的藍色背景）
    purchase_section_start = current_row
    current_row = add_section_header(ws, current_row, "進項", 8, "CCECFF")  # 進項只有8欄
    
    # 進項表頭（2層）
    header_start_row = current_row
    
    # 第1層表頭
    ws.merge_cells(f'A{current_row}:A{current_row+1}')  # 月份 (rowspan=2)
    ws[f'A{current_row}'] = "月份"
    
    ws.merge_cells(f'B{current_row}:C{current_row}')    # 得扣抵載有進項稅額之憑證 (colspan=2)
    ws[f'B{current_row}'] = "得扣抵載有進項稅額之憑證"
    
    ws.merge_cells(f'D{current_row}:E{current_row}')    # 退回折讓 (colspan=2)
    ws[f'D{current_row}'] = "退回折讓"
    
    ws.merge_cells(f'F{current_row}:G{current_row}')    # 合計 (colspan=2)
    ws[f'F{current_row}'] = "合計"
    
    ws.merge_cells(f'H{current_row}:H{current_row+1}')  # 進口貨物金額 (rowspan=2)
    ws[f'H{current_row}'] = "進口免稅貨物金額"
    
    current_row += 1
    
    # 第2層表頭
    ws[f'B{current_row}'] = "進貨及費用"
    ws[f'C{current_row}'] = "固定資產"
    ws[f'D{current_row}'] = "進貨及費用"
    ws[f'E{current_row}'] = "固定資產"
    ws[f'F{current_row}'] = "進貨及費用"
    ws[f'G{current_row}'] = "固定資產"
    
    # 設定進項表頭樣式
    for row in range(header_start_row, current_row + 1):
        for col in range(1, 9):  # 進項只有8欄
            cell = ws.cell(row=row, column=col)
            if cell.value:  # 只設定有值的儲存格
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    current_row += 1
    
    # 填入進項資料
    for record in table_data:
        ws.cell(row=current_row, column=1, value=record.get("month", ""))
        
        # 數字欄位
        purchase_cells_data = [
            (2, record.get("purchaseAndExpense", 0)),
            (3, record.get("fixedAssets", 0)),
            (4, record.get("purchaseReturn", 0)),
            (5, record.get("purchaseReturnAssets", 0)),
            (6, record.get("purchaseTotal", 0)),
            (7, record.get("assetsTotal", 0)),
            (8, record.get("importAmount", 0))
        ]
        
        for col, value in purchase_cells_data:
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
        
        current_row += 1
    
    # 進項合計列
    if table_data:
        ws.cell(row=current_row, column=1, value="合計")
        
        # 計算進項合計
        purchase_totals = {
            "purchaseAndExpense": sum(r.get("purchaseAndExpense", 0) for r in table_data),
            "fixedAssets": sum(r.get("fixedAssets", 0) for r in table_data),
            "purchaseReturn": sum(r.get("purchaseReturn", 0) for r in table_data),
            "purchaseReturnAssets": sum(r.get("purchaseReturnAssets", 0) for r in table_data),
            "purchaseTotal": sum(r.get("purchaseTotal", 0) for r in table_data),
            "assetsTotal": sum(r.get("assetsTotal", 0) for r in table_data),
            "importAmount": sum(r.get("importAmount", 0) for r in table_data)
        }
        
        purchase_total_cells_data = [
            (2, purchase_totals["purchaseAndExpense"]),
            (3, purchase_totals["fixedAssets"]),
            (4, purchase_totals["purchaseReturn"]),
            (5, purchase_totals["purchaseReturnAssets"]),
            (6, purchase_totals["purchaseTotal"]),
            (7, purchase_totals["assetsTotal"]),
            (8, purchase_totals["importAmount"])
        ]
        
        for col, value in purchase_total_cells_data:
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
            cell.font = Font(bold=True)
        
        # 設定合計列邊框
        add_total_border(ws, current_row, 2, 8)
        
        purchase_section_end = current_row
    else:
        purchase_section_end = current_row - 1
    
    # 為進項表格加上黑框（從「進項」標題到合計列）
    add_black_borders(ws, purchase_section_start, purchase_section_end, 1, 8)


def create_withholding_excel(wb: Workbook, version: TaxOcrVersion):
    """
    建立扣繳憑單 Excel
    
    格式：
    - 前3行：公司名稱、申報書類型、年度
    - 第4行：空行
    - 第5行：表頭
    - 資料行：按項目分組，每組後有項目總額列（只在數字欄位有邊框）
    """
    ws = wb.create_sheet("扣繳憑單")
    
    # 設定欄寬（8欄）
    column_widths = [20, 12, 30, 25, 18, 15, 10, 30]
    set_column_widths(ws, column_widths)
    
    # 決定申報書類型名稱
    if version.table_type in ['withholding_expense', 'dividend_expense']:
        doc_type = "各類給付扣繳申報、股利憑單彙總表"
    else:  # withholding_income, dividend_income
        doc_type = "各類收益扣繳憑單、股利憑單彙總表"
    
    # 新增標題行（前3行 + 空行）
    current_row = add_title_rows(ws, version.company_name, doc_type, version.fiscal_year, len(column_widths))
    
    # 設定表頭（8欄）
    headers = ["項目", "底稿索引", "扣繳單位名稱", "所得類別", "各類給付總額", "扣繳稅額", "頁碼", "檔案名稱"]
    header_row = current_row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    current_row += 1
    data_start_row = current_row
    
    # 按項目分組
    table_data = version.table_data or []
    grouped_data = {}
    for record in table_data:
        item_name = record.get("itemName", "未分類")
        if item_name not in grouped_data:
            grouped_data[item_name] = []
        grouped_data[item_name].append(record)
    
    # 項目排序
    item_order = {
        '薪資': 1, '執行業務報酬': 2, '利息': 3, '租賃': 4, '權利金': 5,
        '股利': 6, '競技、競賽及機會中獎': 7, '退職所得': 8, '財產交易': 9,
        '其他所得': 10, '退休金員工自提數': 11
    }
    
    sorted_items = sorted(grouped_data.keys(), key=lambda x: item_order.get(x, 999))
    
    # 填入資料
    for item_name in sorted_items:
        records = grouped_data[item_name]
        
        # 填入該項目的所有記錄
        for record in records:
            ws.cell(row=current_row, column=1, value=record.get("itemName", ""))
            ws.cell(row=current_row, column=2, value=record.get("index", ""))
            ws.cell(row=current_row, column=3, value=record.get("payerName", ""))
            ws.cell(row=current_row, column=4, value=record.get("incomeType", ""))
            
            # 數字欄位設定格式
            total_amount_cell = ws.cell(row=current_row, column=5, value=record.get("totalAmount", 0))
            withholding_tax_cell = ws.cell(row=current_row, column=6, value=record.get("withholdingTax", 0))
            total_amount_cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
            withholding_tax_cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
            
            # 🆕 新增頁碼和檔案名稱（頁碼支援股利憑單的 "頁碼-序號" 格式）
            ws.cell(row=current_row, column=7, value=format_page_number(record))
            ws.cell(row=current_row, column=8, value=record.get("fileName", ""))
            
            current_row += 1
        
        # 每個項目都要有總額列（不管幾筆記錄）
        # 計算總額
        total_amount = sum(record.get("totalAmount", 0) for record in records)
        total_withholding = sum(record.get("withholdingTax", 0) for record in records)
        
        # 前4欄留空
        ws.cell(row=current_row, column=1, value="")
        ws.cell(row=current_row, column=2, value="")
        ws.cell(row=current_row, column=3, value="")
        ws.cell(row=current_row, column=4, value="")
        
        # 填入總額（只在數字欄位）
        total_amount_cell = ws.cell(row=current_row, column=5, value=total_amount)
        total_withholding_cell = ws.cell(row=current_row, column=6, value=total_withholding)
        
        # 設定數字格式
        total_amount_cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
        total_withholding_cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
        
        # 🆕 第7和第8欄留空（頁碼和檔案名稱）
        ws.cell(row=current_row, column=7, value="")
        ws.cell(row=current_row, column=8, value="")
        
        # 設定邊框（只在數字欄位：第5和第6欄）
        add_subtotal_border(ws, current_row, 5, 6)
        
        current_row += 1
        
        # 總額列後空一行
        current_row += 1
    
    # 只為表頭加上黑框（8欄）
    add_black_borders(ws, header_row, header_row, 1, 8)


def aggregate_type3_results(jobs: list[Job]) -> dict:
    """
    聚合 TYPE3 (憑單) 的結果

    Args:
        jobs: Job 物件列表

    Returns:
        聚合後的資料

    Note:
        TYPE3 同一 PDF 內的所有頁面需要加總
        但如果是不同 PDF，則保留為列表（或根據需求加總）
    """
    # TODO: 確認跨 PDF 的聚合邏輯
    # 目前先簡單列出所有結果
    results = []
    for job in jobs:
        if job.result_json:
            results.append(job.result_json)
    return {"資料": results}
